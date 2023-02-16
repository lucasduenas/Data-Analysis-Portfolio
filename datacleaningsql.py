#openpyxl------------------------------------------------------------------------
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
#os-------------------------------------------------------------------------------
import win32com.client as win32
import sys
import os
#MySQL--------------------
import mysql.connector


#Connect to MySQL database
conn = mysql.connector.connect(
    host='localhost',
    user='root',
    passwd='root',
    database='PortfolioProject'
)

cur=conn.cursor()

cur.execute('CREATE DATABASE IF NOT EXISTS PortfolioProject')



def resource_path(relative_path):
#Get absolute path to resource, works for dev and for PyInstaller
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


try:
    nashville_wb=load_workbook(resource_path('NashvilleHousingData.xlsx'))

except:
    nashville_wb=load_workbook('NashvilleHousingData.xlsx')

data_ws=nashville_wb.active

##Create Data table and populate titles
wb_titles=[str(cell.value).strip() for col in data_ws['A1':'S1'] for cell in col]
""" 

cur.execute(f'CREATE TABLE IF NOT EXISTS Data ({wb_titles[0]} TEXT)')
conn.commit()

for title in wb_titles[1:]:
    print(title)
    cur.execute(f'''ALTER TABLE Data ADD {title} TEXT''')

    conn.commit() """


for row in data_ws.iter_rows(min_row=2, 
    max_col=len(wb_titles), 
    values_only=True):
    cur.execute(f'''INSERT INTO data VALUES (
        "{str(row[0]).strip()}",
        "{str(row[1]).strip()}",
        "{str(row[2]).strip()}",
        "{str(row[3]).strip()}",
        "{str(row[4]).strip()}",
        "{str(row[5]).strip()}",
        "{str(row[6]).strip()}",
        "{str(row[7]).strip()}",
        "{str(row[8]).strip()}",
        "{str(row[9]).strip()}",
        "{str(row[10]).strip()}",
        "{str(row[11]).strip()}",
        "{str(row[12]).strip()}",
        "{str(row[13]).strip()}",
        "{str(row[14]).strip()}",
        "{str(row[15]).strip()}",
        "{str(row[16]).strip()}",
        "{str(row[17]).strip()}",
        "{str(row[18]).strip()}"      
        )'''
        
        )
conn.commit()
conn.close()